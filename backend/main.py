"""
backend/main.py

Entry point FastAPI untuk QS BOQ AI Extension API.
Menyediakan 6 endpoint:

1. POST /api/detect-template  — Deteksi format template Excel
2. POST /api/extract-file     — Ekstrak dimensi dari gambar kerja
3. POST /api/extract-pdf      — Alias backward-compat ke /api/extract-file
4. POST /api/generate-boq     — Generate file Excel BOQ
5. POST /api/chat             — QS Assistant chat
6. POST /api/byok/validate    — Validasi BYOK API key
7. POST /api/byok/save        — Simpan BYOK API key
8. POST /api/byok/delete      — Hapus BYOK API key user
9. POST /api/compute-boq      — Compute BOQ formula & return JSON (dengan template file)
10. POST /api/compute-boq-from-json — Compute BOQ formula dari JSON headers (untuk Excel Add-in)
"""

import json
import os
import tempfile
from pathlib import Path
from dotenv import load_dotenv

from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles

env_path = Path(__file__).resolve().parent / ".env"
load_dotenv(dotenv_path=env_path)

app = FastAPI(title="QS BOQ AI Extension API")

EXCEL_ADDIN_DIR = Path(__file__).resolve().parent.parent / "excel-addin"
if EXCEL_ADDIN_DIR.exists():
    app.mount("/static/excel-addin", StaticFiles(directory=str(EXCEL_ADDIN_DIR)), name="excel_addin")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

byok_manager = None


@app.on_event("startup")
async def startup():
    global byok_manager
    try:
        from modules.byok_manager import BYOKManager
        byok_manager = BYOKManager()
    except Exception as e:
        print(f"[WARN] BYOK tidak aktif: {e}")
        byok_manager = None


SUPPORTED_EXT: set[str] = {
    ".pdf", ".jpg", ".jpeg", ".png", ".webp",
    ".tiff", ".tif", ".bmp", ".heic", ".heif", ".dxf",
}


@app.get("/api/health")
async def health():
    """Health check endpoint untuk monitoring."""
    return {
        "status": "ok",
        "service": "QS BOQ AI Extension API",
        "byok_active": byok_manager is not None,
    }


@app.get("/api/addin-info")
async def addin_info():
    """Info untuk install Excel Add-in / Google Sheets Add-in."""
    manifest_url = f"{os.getenv('PUBLIC_URL', 'http://localhost:8000')}/static/excel-addin/manifest.xml"
    return {
        "status": "ok",
        "excel": {
            "name": "QS BOQ AI",
            "manifest_url": manifest_url,
            "install": "Excel -> Insert -> Get Add-ins -> From URL",
        },
        "google_sheets": {
            "folder": "google-sheets-addin/",
            "files": ["Code.gs", "Sidebar.html"],
            "install": "Extensions -> Apps Script -> copy Code.gs + Sidebar.html",
        },
        "backend": {
            "url": os.getenv("9ROUTER_BASE_URL", "http://localhost:20128/v1"),
            "model": os.getenv("9ROUTER_MODEL", "groq/llama-3.3-70b-versatile"),
        },
    }


@app.post("/api/detect-template")
async def detect_template(
    file: UploadFile = File(...),
    user_id: str = Form(""),
):
    """Baca template Excel user dan return mapping kolom."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name
    try:
        from modules.template_detector import detect_template_format
        result = detect_template_format(tmp_path)
        return {"status": "ok", "data": result}
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


@app.post("/api/extract-file")
async def extract_file(
    file: UploadFile = File(...),
    user_id: str = Form(""),
    scale: float = Form(None),
):
    """Ekstrak dimensi dari gambar kerja. Mendukung PDF, JPG/PNG/WEBP/TIFF/BMP/HEIC, DXF."""
    original_filename = file.filename or "upload.pdf"
    ext = Path(original_filename).suffix.lower()
    if ext not in SUPPORTED_EXT:
        raise HTTPException(
            400,
            f"Format {ext} tidak didukung. "
            f"Format yang didukung: {', '.join(sorted(SUPPORTED_EXT))}",
        )

    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name

    try:
        from modules.llm_router import LLMRouter
        from modules.file_router import route_to_reader

        byok_key = byok_manager.get_key(user_id) if byok_manager and user_id else None
        router = LLMRouter(byok_key=byok_key)
        result = route_to_reader(tmp_path, llm_router=router, scale=scale)
        return {"status": "ok", "data": result, "original_filename": original_filename}
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


@app.post("/api/extract-pdf")
async def extract_pdf_legacy(
    file: UploadFile = File(...),
    user_id: str = Form(""),
    scale: float = Form(None),
):
    """Alias ke /api/extract-file untuk backward compatibility."""
    return await extract_file(file=file, user_id=user_id, scale=scale)


@app.post("/api/generate-boq")
async def generate_boq(
    template_file: UploadFile = File(...),
    dimensions_json: str = Form(...),
    user_id: str = Form(""),
):
    """Generate file Excel BOQ dengan rumus hidup berdasarkan template user."""
    try:
        dimensions = json.loads(dimensions_json)
    except json.JSONDecodeError:
        raise HTTPException(400, "dimensions_json harus berupa JSON array yang valid")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(await template_file.read())
        template_path = tmp.name

    base, ext = os.path.splitext(template_path)
    output_path = f"{base}_boq_output{ext}"

    try:
        from modules.template_detector import detect_template_format
        from modules.formula_engine import build_formula
        from modules.excel_writer import write_boq_to_excel

        mapping = detect_template_format(template_path)
        boq_data = []

        for i, item in enumerate(dimensions.get("items", [])):
            formula_data = build_formula(
                item=item,
                mapping=mapping["mapping"],
                row=mapping["data_start_row"] + i,
                case=mapping["case"],
            )
            boq_data.append(formula_data)

        write_boq_to_excel(template_path, boq_data, mapping["mapping"], output_path)

        from fastapi.responses import FileResponse

        def cleanup():
            if os.path.exists(output_path):
                os.unlink(output_path)

        return FileResponse(
            output_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="BOQ_Output.xlsx",
            background=None,
        )
    finally:
        if os.path.exists(template_path):
            os.unlink(template_path)


@app.post("/api/compute-boq")
async def compute_boq(
    template_file: UploadFile = File(...),
    dimensions_json: str = Form(...),
    user_id: str = Form(""),
):
    """Compute BOQ formulas dan return JSON cell data (untuk GAS/Excel Add-in)."""
    try:
        dimensions = json.loads(dimensions_json)
    except json.JSONDecodeError:
        raise HTTPException(400, "dimensions_json harus berupa JSON array yang valid")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(await template_file.read())
        template_path = tmp.name

    try:
        from modules.template_detector import detect_template_format
        from modules.formula_engine import build_formula

        mapping = detect_template_format(template_path)
        boq_data = []

        for i, item in enumerate(dimensions.get("items", [])):
            formula_data = build_formula(
                item=item,
                mapping=mapping["mapping"],
                row=mapping["data_start_row"] + i,
                case=mapping["case"],
            )
            boq_data.append(formula_data)

        return {
            "status": "ok",
            "data": {
                "items": boq_data,
                "mapping": mapping["mapping"],
                "data_start_row": mapping["data_start_row"],
            },
        }
    finally:
        if os.path.exists(template_path):
            os.unlink(template_path)


@app.post("/api/compute-boq-from-json")
async def compute_boq_from_json(request: dict):
    """Compute BOQ formulas dari JSON headers (untuk Excel Add-in tanpa upload template)."""
    dimensions = request.get("dimensions")
    headers = request.get("headers")
    column_map = request.get("column_map")
    if not dimensions or not headers:
        raise HTTPException(400, "dimensions dan headers required")

    from modules.template_detector import detect_template_format
    from modules.formula_engine import build_formula
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    for col, header in enumerate(headers, 1):
        ws.cell(1, col).value = header
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    wb.close()
    tmp.close()

    try:
        mapping = detect_template_format(tmp.name)

        if column_map:
            for category, cat_info in mapping["mapping"].items():
                if cat_info and cat_info.get("nama_kolom") in column_map:
                    actual_col = column_map[cat_info["nama_kolom"]]
                    cat_info["col_num"] = actual_col
                    cat_info["col_index"] = get_column_letter(actual_col)

        boq_data = []
        for i, item in enumerate(dimensions.get("items", [])):
            formula_data = build_formula(
                item=item,
                mapping=mapping["mapping"],
                row=mapping["data_start_row"] + i,
                case=mapping["case"],
            )
            boq_data.append(formula_data)

        return {
            "status": "ok",
            "data": {
                "items": boq_data,
                "mapping": mapping["mapping"],
                "data_start_row": mapping["data_start_row"],
            },
        }
    finally:
        if os.path.exists(tmp.name):
            os.unlink(tmp.name)


@app.post("/api/chat")
async def chat(request: dict):
    """QS Assistant chat. Body: {user_message, boq_state, template_mapping, file_names, user_id}"""
    from modules.llm_router import LLMRouter
    from modules.chat_handler import ChatHandler

    user_message = request.get("user_message")
    if not user_message:
        raise HTTPException(400, "user_message required")
    byok_key = byok_manager.get_key(request.get("user_id", "")) if byok_manager else None
    router = LLMRouter(byok_key=byok_key)
    handler = ChatHandler(router)
    result = handler.chat(
        user_message=user_message,
        boq_state=request.get("boq_state", []),
        template_mapping=request.get("template_mapping", {}),
        file_names=request.get("file_names", {}),
    )
    return {"status": "ok", "data": result}


@app.post("/api/byok/validate")
async def validate_byok(request: dict):
    """Validasi API key Gemini milik user."""
    if byok_manager is None:
        raise HTTPException(503, "BYOK service tidak tersedia")
    api_key = request.get("api_key")
    if not api_key:
        raise HTTPException(400, "api_key required")
    result = byok_manager.validate_key(api_key)
    return {"status": "ok", "data": result}


@app.post("/api/byok/save")
async def save_byok(request: dict):
    """Simpan API key user (terenkripsi)."""
    if byok_manager is None:
        raise HTTPException(503, "BYOK service tidak tersedia")
    user_id = request.get("user_id")
    api_key = request.get("api_key")
    if not user_id or not api_key:
        raise HTTPException(400, "user_id dan api_key required")
    success = byok_manager.save_key(user_id, api_key)
    return {"status": "ok" if success else "error", "data": {"saved": success}}


@app.post("/api/byok/delete")
async def delete_byok(request: dict):
    """Hapus API key user (terenkripsi)."""
    if byok_manager is None:
        raise HTTPException(503, "BYOK service tidak tersedia")
    user_id = request.get("user_id")
    if not user_id:
        raise HTTPException(400, "user_id required")
    deleted = byok_manager.delete_key(user_id)
    return {"status": "ok", "data": {"deleted": deleted}}
