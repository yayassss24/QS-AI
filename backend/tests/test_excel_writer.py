"""Tests for excel_writer.py"""

import os
import tempfile
from modules.excel_writer import write_boq_to_excel


def test_write_boq_to_excel_creates_file(temp_xlsx, sample_mapping):
    boq_data = [
        {
            "row": 2,
            "cells": {
                "C2": {"value": 12.5, "type": "dimension", "color": "#EEEDFE"},
                "D2": {"value": 0.8, "type": "dimension", "color": "#EEEDFE"},
                "E2": {"value": 1.2, "type": "dimension", "color": "#EEEDFE"},
                "G2": {"value": "=C2*D2*E2", "type": "formula"},
            },
            "comments": {"C2": "Sumber: Gambar hal.2"},
        }
    ]

    out = tempfile.mktemp(suffix=".xlsx")
    try:
        result = write_boq_to_excel(temp_xlsx, boq_data, sample_mapping, out)
        assert result["status"] == "ok"
        assert os.path.exists(out)

        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb.active
        assert ws["C2"].value == 12.5
        assert ws["G2"].value == "=C2*D2*E2"
        wb.close()
    finally:
        if os.path.exists(out):
            os.unlink(out)


def test_formula_written_as_live_formula(temp_xlsx, sample_mapping):
    boq_data = [
        {
            "row": 3,
            "cells": {
                "G3": {"value": "=C3*D3*E3", "type": "formula"},
            },
            "comments": {},
        }
    ]

    out = tempfile.mktemp(suffix=".xlsx")
    try:
        write_boq_to_excel(temp_xlsx, boq_data, sample_mapping, out)
        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb.active
        val = ws["G3"].value
        assert isinstance(val, str) and val.startswith("=")
        wb.close()
    finally:
        if os.path.exists(out):
            os.unlink(out)


def test_empty_cell_set_to_none(temp_xlsx, sample_mapping):
    boq_data = [
        {
            "row": 2,
            "cells": {
                "E2": {"value": None, "type": "empty"},
            },
            "comments": {},
        }
    ]

    out = tempfile.mktemp(suffix=".xlsx")
    try:
        write_boq_to_excel(temp_xlsx, boq_data, sample_mapping, out)
        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb.active
        assert ws["E2"].value is None
        wb.close()
    finally:
        if os.path.exists(out):
            os.unlink(out)
