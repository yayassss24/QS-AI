"""
Shared fixtures for all tests.
Creates temp files, mock objects, and sample data.
"""

import os
import sys
from pathlib import Path

# Make backend/ importable so `from modules.xxx` and `from utils.xxx` work
BACKEND_DIR = str(Path(__file__).resolve().parent.parent)
if BACKEND_DIR not in sys.path:
    sys.path.insert(0, BACKEND_DIR)

# Set default keys for tests that don't mock them
if "BYOK_ENCRYPTION_KEY" not in os.environ:
    os.environ["BYOK_ENCRYPTION_KEY"] = "dGVzdC1mZXJuZXQta2V5LTEyMzQ1Njc4OTAxMjM0NTY3ODkwMTIzNA=="
if "9ROUTER_API_KEY" not in os.environ:
    os.environ["9ROUTER_API_KEY"] = "sk-test-key-for-tests"

# Disable startup events for main.py tests
os.environ.setdefault("TESTING", "1")

import json
import os
import tempfile
from pathlib import Path
from unittest.mock import MagicMock

import pytest


@pytest.fixture
def sample_items():
    return [
        {
            "nama_item": "Galian Tanah Pondasi",
            "P": 12.5,
            "L": 0.8,
            "T": 1.2,
            "satuan": "m³",
            "confidence": 0.92,
        },
        {
            "nama_item": "Plesteran Dinding",
            "P": 45.0,
            "L": 3.0,
            "T": None,
            "satuan": "m²",
            "confidence": 0.85,
        },
        {
            "nama_item": "Besi Beton",
            "P": 6.0,
            "L": None,
            "T": None,
            "satuan": "m",
            "confidence": 0.45,
        },
    ]


@pytest.fixture
def sample_mapping():
    return {
        "dimensi_panjang": {"nama_kolom": "Pjg", "col_index": "C", "col_num": 3},
        "dimensi_lebar": {"nama_kolom": "Lbr", "col_index": "D", "col_num": 4},
        "dimensi_tinggi": {"nama_kolom": "T", "col_index": "E", "col_num": 5},
        "kolom_rumus": {"nama_kolom": "Rumus", "col_index": "F", "col_num": 6},
        "kolom_volume": {"nama_kolom": "Volume", "col_index": "G", "col_num": 7},
        "kolom_satuan": {"nama_kolom": "Sat", "col_index": "H", "col_num": 8},
        "kolom_uraian": {"nama_kolom": "Uraian", "col_index": "B", "col_num": 2},
        "kolom_koefisien": None,
    }


@pytest.fixture
def sample_dimension_data():
    return {
        "status": "ok",
        "source": "pymupdf",
        "items": [
            {
                "nama_item": "Galian Tanah",
                "P": 12.5,
                "L": 0.8,
                "T": 1.2,
                "satuan": "m³",
                "confidence": 0.92,
            },
            {
                "nama_item": "Pasangan Batu Kali",
                "P": 12.5,
                "L": 0.3,
                "T": 0.6,
                "satuan": "m³",
                "confidence": 0.78,
            },
        ],
        "items_flagged": [
            {
                "nama_item": "Aanstamping",
                "P": 12.5,
                "L": 0.6,
                "T": None,
                "confidence": 0.45,
                "alasan_flag": "Dimensi T tidak terdeteksi",
            }
        ],
    }


@pytest.fixture
def mock_llm_router():
    router = MagicMock()
    router.call.return_value = json.dumps([
        {
            "nama_item": "Galian Tanah Pondasi",
            "P": 12.5,
            "L": 0.8,
            "T": 1.2,
            "satuan": "m³",
            "confidence": 0.92,
        }
    ])
    return router


@pytest.fixture
def temp_xlsx():
    """Create a minimal valid .xlsx file for template tests."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"
    headers = ["No", "Uraian Pekerjaan", "Pjg", "Lbr", "T", "Rumus", "Volume", "Sat"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    ws.cell(2, 1, 1)
    ws.cell(2, 2, "Galian Tanah")
    ws.cell(2, 3, 12.5)
    ws.cell(2, 4, 0.8)
    ws.cell(2, 5, 1.2)
    ws.cell(2, 7, 12.0)
    ws.cell(2, 8, "m³")
    ws.cell(3, 1, 2)
    ws.cell(3, 2, "Pasangan Batu Kali")
    ws.cell(3, 3, 12.5)
    ws.cell(3, 4, 0.3)
    ws.cell(3, 5, 0.6)

    path = tempfile.mktemp(suffix=".xlsx")
    wb.save(path)
    wb.close()
    yield path
    if os.path.exists(path):
        os.unlink(path)


@pytest.fixture
def temp_image():
    """Create a minimal test PNG image."""
    from PIL import Image
    path = tempfile.mktemp(suffix=".png")
    img = Image.new("RGB", (100, 100), color=(255, 255, 255))
    img.save(path)
    yield path
    if os.path.exists(path):
        os.unlink(path)


@pytest.fixture
def fernet_key():
    from cryptography.fernet import Fernet
    return Fernet.generate_key().decode()
