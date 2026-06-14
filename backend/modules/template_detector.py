"""
modules/template_detector.py

Deteksi format kolom template Excel BOQ menggunakan kombinasi exact match
dan semantic similarity. Mengklasifikasikan template ke dalam 3 case:

Case A: Ada P+L+T + kolom Rumus + kolom Volume
Case B: Ada P+L+T + TIDAK ada kolom Rumus + ada kolom Volume
Case C: TIDAK ada P/L/T + hanya ada kolom Volume/Jumlah
"""

from typing import Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

SEMANTIC_MAP: dict[str, list[str]] = {
    "dimensi_panjang": [
        "panjang", "p", "length", "pjg", "p (m)", "panjang (m)",
        "p(m)", "pan", "long",
    ],
    "dimensi_lebar": [
        "lebar", "l", "width", "lbr", "l (m)", "lebar (m)",
        "l(m)", "lbr", "wide",
    ],
    "dimensi_tinggi": [
        "tinggi", "t", "height", "tggi", "t (m)", "tinggi (m)",
        "t(m)", "h", "h (m)", "kedalaman", "depth", "tebal", "thick",
    ],
    "kolom_volume": [
        "volume", "vol", "kubikasi", "jumlah", "qty", "quantity",
        "kuantitas", "m3", "m²", "m2", "hasil",
    ],
    "kolom_rumus": [
        "rumus", "formula", "keterangan", "ket", "perhitungan",
        "calculation", "detail", "cara hitung",
    ],
    "kolom_satuan": [
        "satuan", "sat", "unit", "uom", "units",
    ],
    "kolom_uraian": [
        "uraian", "pekerjaan", "uraian pekerjaan", "item",
        "description", "desc", "keterangan pekerjaan", "nama",
    ],
    "kolom_koefisien": [
        "koefisien", "koef", "coefficient", "faktor", "factor",
        "k", "n", "jumlah item", "buah", "pcs",
    ],
}


def _find_header_row(ws: Worksheet) -> int:
    """
    Cari baris header dengan heuristik:
    - Baris pertama yang punya 3+ cell terisi berurutan
    - Atau baris yang mengandung kata kunci dimensi
    """
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    for row_num in range(1, min(10, max_row + 1)):
        row_values = [ws.cell(row_num, c).value for c in range(1, max_col + 1)]
        filled = [v for v in row_values if v is not None]
        if len(filled) >= 3:
            row_str = " ".join(str(v).lower() for v in filled)
            keywords = ["panjang", "lebar", "volume", "uraian", "satuan", "p", "l"]
            if any(kw in row_str for kw in keywords):
                return row_num
    return 1


def _col_letter(col_num: int) -> str:
    """Convert 1-based column number to Excel letter (A, B, ..., Z, AA, ...)."""
    letter = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        letter = chr(65 + remainder) + letter
    return letter


def _semantic_match(header_value) -> Optional[str]:
    """
    Cocokkan nama kolom ke kategori menggunakan exact + fuzzy match.
    Return kategori (dimensi_panjang, dll) atau None jika tidak cocok.
    """
    if header_value is None:
        return None
    val = str(header_value).lower().strip()

    for kategori, keywords in SEMANTIC_MAP.items():
        if val in keywords:
            return kategori

    # Partial match: minimal 3 karakter untuk kw in val (hindari false positive "p","l","t")
    for kategori, keywords in SEMANTIC_MAP.items():
        for kw in keywords:
            if len(kw) >= 3 and kw in val:
                return kategori
            if val in kw:
                return kategori
    return None


def detect_template_format(file_path: str) -> dict:
    """
    Baca template Excel user dan deteksi struktur kolom.

    Args:
        file_path: Path ke file Excel (.xlsx)

    Returns:
        dict dengan status, case, header_row, data_start_row, mapping, dll
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    header_row = _find_header_row(ws)
    data_start_row = header_row + 1

    ws_max_row = ws.max_row or 0
    max_col_ws = ws.max_column or 0
    headers: dict[int, str] = {}
    for col in range(1, max_col_ws + 1):
        val = ws.cell(header_row, col).value
        if val is not None:
            headers[col] = str(val).strip()

    mapping: dict[str, Optional[dict]] = {
        "dimensi_panjang": None,
        "dimensi_lebar": None,
        "dimensi_tinggi": None,
        "kolom_rumus": None,
        "kolom_volume": None,
        "kolom_satuan": None,
        "kolom_uraian": None,
        "kolom_koefisien": None,
    }

    for col_num, header_text in headers.items():
        kategori = _semantic_match(header_text)
        if kategori and kategori in mapping:
            mapping[kategori] = {
                "nama_kolom": header_text,
                "col_index": _col_letter(col_num),
                "col_num": col_num,
            }

    has_p = mapping["dimensi_panjang"] is not None
    has_l = mapping["dimensi_lebar"] is not None
    has_t = mapping["dimensi_tinggi"] is not None
    has_rumus = mapping["kolom_rumus"] is not None
    has_volume = mapping["kolom_volume"] is not None

    if has_p and has_l and has_t and has_rumus and has_volume:
        case = "A"
    elif has_p and has_l and has_t and has_volume:
        case = "B"
    else:
        case = "C"

    merged_ranges = []
    has_merged = False
    if ws.merged_cells.ranges:
        has_merged = True
        merged_ranges = [str(mr) for mr in ws.merged_cells.ranges]

    section_rows: list[int] = []
    for row_num in range(data_start_row, ws_max_row + 1):
        row_val = ws.cell(row_num, 1).value
        if row_val and isinstance(row_val, str) and (
            "pekerjaan" in row_val.lower()
            or row_val.strip().startswith("I.")
            or row_val.strip().startswith("II.")
            or row_val.strip().startswith("III.")
        ):
            section_rows.append(row_num)

    wb.close()

    return {
        "status": "ok",
        "case": case,
        "header_row": header_row,
        "data_start_row": data_start_row,
        "mapping": mapping,
        "has_merged_cells": has_merged,
        "merged_ranges": merged_ranges,
        "section_rows": section_rows,
        "total_rows": ws.max_row or 0,
    }
