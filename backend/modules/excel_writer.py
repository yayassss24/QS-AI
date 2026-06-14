"""
modules/excel_writer.py

Tulis data BOQ ke file Excel berdasarkan template user.
Preserves semua formatting asli template - hanya mengisi data di cell.
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment

COLOR_AI_INPUT = "EEEDFE"
COLOR_FLAG_WARN = "FAEEDA"
COLOR_FORMULA = None
COLOR_SUBTOTAL = "EAF3DE"


def write_boq_to_excel(
    template_path: str,
    boq_data: list[dict],
    mapping: dict,
    output_path: str,
) -> dict:
    """
    Tulis data BOQ ke file Excel berdasarkan template user.

    Args:
        template_path: Path ke template Excel yang akan diisi
        boq_data: List output dari formula_engine per item
        mapping: Output mapping dari template_detector
        output_path: Path untuk menyimpan file Excel hasil

    Returns:
        dict dengan status dan output_path
    """
    wb = load_workbook(template_path)
    ws = wb.active

    for item_data in boq_data:
        row = item_data["row"]
        for cell_addr, cell_info in item_data["cells"].items():
            cell = ws[cell_addr]

            if cell_info["type"] == "formula":
                cell.value = cell_info["value"]
            elif cell_info["type"] == "dimension":
                cell.value = cell_info["value"]
                cell.fill = PatternFill("solid", fgColor=COLOR_AI_INPUT)
            elif cell_info["type"] == "angka":
                cell.value = cell_info["value"]
            elif cell_info["type"] == "empty":
                cell.value = None

        for cell_addr, comment_text in item_data.get("comments", {}).items():
            cell = ws[cell_addr]
            cell.comment = Comment(comment_text, "QS AI Assistant")
            if "⚠️" in comment_text:
                cell.fill = PatternFill("solid", fgColor=COLOR_FLAG_WARN)

    wb.save(output_path)
    wb.close()
    return {"status": "ok", "output_path": output_path}
